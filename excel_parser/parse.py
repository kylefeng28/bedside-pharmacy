import openpyxl as xl
import sys
import json
from functools import reduce
def parse_worksheet(ws):
    data = {
        'name': ws.title,
        'drugs': []
    }

    # Find column names
    column_names = ws['1']

    #print(column_names)
    #    column_names.append(name.replace('/', ''))
    column_names = list(filter(lambda cell: cell.value != None, column_names))
    column_indices = [ c.column for c in column_names ]

    # Find row names
    # read-only mode has no iter_col, so do it like this
    row_names = [ x[0] for x in ws.iter_rows(min_row=0, max_row=500, min_col=0, max_col=0) ]
    row_names = list(filter(lambda cell: cell.value != None, row_names))
    row_indices = [ r.row for r in row_names ]

    # Find where last row ends based on bottom border
    last_row = row_names[-1].row
    while last_row <= 500: # set max bound
        cell = ws.cell(column=1, row=last_row)
        if cell.border.bottom.color is not None: # found
            break
        last_row += 1
    row_indices.append(last_row + 1) # append 1 and add to list

    # Insert data
    # Iterate by rows (drug name)
    for (row_i, row_name) in enumerate(row_names):
        drug_data = {}
        drug = {
            'name': row_name.value,
            '_coordinate': row_name.coordinate,
            'data': drug_data
        }
        data['drugs'].append(drug)

        # Iterate by columns
        # ex. DOSE, ONSET/DURATION, METABOLISM/EXCRETION, WARNINGS
        for (col_i, col_name) in enumerate(column_names):
            col_name_value = col_name.value.replace('/','|')
            drug_data[col_name_value] = {}
            #print(drug_data[col_name.value])
            col = column_indices[col_i]
            min_row = row_indices[row_i]
            max_row = row_indices[row_i+1]

            # Iterate by "subcolumns", which are bolded items within a column
            subcolumn_name = None
            for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=col, max_col=col):
                if r[0].value is not None:
                    if r[0].font.bold:
                        subcolumn_name = r[0].value
                        drug_data[col_name_value][subcolumn_name] = ''
                    else:
                        drug_data[col_name_value][subcolumn_name] += r[0].value + '\n'
                   # print(r[0].value)

            #print(drug_data[col_name.value])
    return data

if __name__ == '__main__':
    # Parse arguments
    if len(sys.argv) != 2:
        print("usage: parse.py filename")
        sys.exit()

    filename = sys.argv[1]
    data = {}

    # Load workbook
    wb = xl.load_workbook(filename, read_only=True)

    # Loop through worksheets
    # for ws in wb:
    #     parse_worksheet(wb, data)
    data['classes'] = []
    data['classes'].append(parse_worksheet(wb['BENZODIAZEPINES']))

    # Print JSON
    with open('data.json', 'w') as outfile:
        json.dump(data, outfile, indent = 4)
